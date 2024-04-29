from tkinter import *
from database import Produtos, Vendas
from datetime import datetime
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws["A1"] = "item"
ws["B1"] = "preço"
ws["C1"] = "quantidade"
ws["D1"] = "Total"
ws["E1"] = "vendador"
ws["F1"] = "data"



def realizar_venda():
    itemv = item.get()
    quantidadev = int(quantidade.get())
    atual_item = -1
    for item_auth in Produtos.select():
        atual_item += 1
        if item_auth.item == itemv:
            if item_auth.quantidade >= quantidadev:
                if item_auth.quantidade == 0:
                    erro2["text"] = "Produto Indisponivel"
                else:
                    item_auth.quantidade = item_auth.quantidade - quantidadev
                    item_auth.save()
                    data = datetime.today().strftime("%d-%m-%Y")

                    Vendas.create(
                        item=item_auth.item,
                        preco=item_auth.preco,
                        quantidade=quantidadev,
                        vendedor = user,
                        data=data,
                    )
                    gerenc.destroy()
                    venda.destroy()
                    gerenc_interface(user)
            else:
                erro2["text"] = "Quantidade invalida"
        else:
            erro["text"] = "Produto não encontrado"


def painel_vendas():
    global item, quantidade, erro, erro2, venda
    venda = Tk()
    item = Entry(venda)
    item.grid(column=0, row=1)
    quantidade = Entry(venda)
    quantidade.grid(column=1, row=1)
    erro = Label(venda, text="")
    erro2 = Label(venda, text="")
    erro.grid(column=0, row=3)
    erro2.grid(column=1, row=3)

    Label(venda, text="Produto").grid(column=0, row=0)
    Label(venda, text="Quantidade").grid(column=1, row=0)
    Button(venda, text="Atribuir Venda", command=realizar_venda).grid(column=3, row=1)

    venda.mainloop()

def fechamento():
    dados = []

    total_count = 0


    for dado in Vendas.select():
        total_count += float(dado.quantidade * dado.preco)
        dados.append((dado.item, dado.preco, dado.quantidade, total_count,dado.vendedor, dado.data))
        total_count -= float(dado.quantidade * dado.preco)

    for linha in dados:
        ws.append(linha)

    data = datetime.today().strftime("%d-%m-%Y")
    wb.save(f"Fechamento-{data}.xlsx")

    for delete in Vendas.select():
        delete.delete_instance()
    historico.destroy()


def Historico():
    global historico
    historico = Tk()

    Button(historico, text="Fechamento de Caixa", command=fechamento).grid(column=2, row=0)

    ycount = 1
    Label(historico, text="item").grid(column=0, row=1)
    Label(historico, text="preço").grid(column=1, row=1)
    Label(historico, text="quantidade").grid(column=2, row=1)
    Label(historico, text="vendedor").grid(column=3, row=1)
    Label(historico, text="data").grid(column=4, row=1)



    for prod in Vendas.select():
        ycount += 1
        Label(historico, text=prod.item).grid(column=0, row=ycount, pady=10)
        Label(historico, text=prod.preco).grid(column=1, row=ycount, pady=10)
        Label(historico, text=prod.quantidade).grid(column=2, row=ycount, pady=10)
        Label(historico, text=prod.vendedor).grid(column=3, row=ycount, pady=10)
        Label(historico, text=prod.data).grid(column=4, row=ycount, pady=10)


    historico.mainloop()

def gerenc_interface(user_logado):
    global gerenc, user
    gerenc = Tk()
    user = user_logado
    gerenc.title("Gerenciamento")
    gerenc.geometry("500x300")
    ycount = 0.4
    label_produtos = Label(gerenc, text="Produtos: ")
    label_produtos.place(relx=0.27, rely=0.3, anchor=CENTER)

    label_precos = Label(gerenc, text="Preços: ")
    label_precos.place(relx=0.47, rely=0.3, anchor=CENTER)

    label_quantidade = Label(gerenc, text="Quantidade: ")
    label_quantidade.place(relx=0.67, rely=0.3, anchor=CENTER)

    Button(gerenc, text="Realizar Venda", command=painel_vendas).pack()
    Button(gerenc, text="Historico de Vendas", command=Historico).pack()

    for prod in Produtos.select():
        Label(gerenc, text=prod.item).place(relx=0.27, rely=ycount, anchor=CENTER)
        Label(gerenc, text=prod.preco).place(relx=0.47, rely=ycount, anchor=CENTER)
        Label(gerenc, text=prod.quantidade).place(relx=0.67, rely=ycount, anchor=CENTER)
        ycount += 0.1

    gerenc.mainloop()
